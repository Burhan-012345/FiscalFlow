import os
import secrets
import logging
from datetime import datetime, timedelta
import traceback
from flask import url_for, render_template, current_app
from flask_mail import Message
from itsdangerous import URLSafeTimedSerializer
from PIL import Image
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import json
from werkzeug.utils import secure_filename

# Set up logging
logger = logging.getLogger(__name__)

class SecurityUtils:
    def __init__(self, app=None):
        self.app = app
        if app:
            self.serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

    def generate_token(self, data, salt=None):
        """Generate a secure token"""
        try:
            return self.serializer.dumps(data, salt=salt)
        except Exception as e:
            logger.error(f"Token generation failed: {e}")
            return None

    def verify_token(self, token, max_age=3600, salt=None):
        """Verify and decode a token"""
        try:
            data = self.serializer.loads(token, max_age=max_age, salt=salt)
            return data
        except Exception as e:
            logger.warning(f"Token verification failed: {e}")
            return None

class EmailUtils:
    def __init__(self, mail, app):
        self.mail = mail
        self.app = app
        self.logger = logging.getLogger(__name__)

    def send_otp_email(self, email, otp_code):
        """Send OTP verification email with comprehensive error handling"""
        try:
            subject = "FiscalFlow - Email Verification Code"

            # Fallback HTML content
            html_content = self._create_otp_email_fallback(otp_code)

            # Try to use template if it exists
            try:
                html_content = render_template('email/otp.html',
                                             otp_code=otp_code,
                                             app_url=self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000'))
            except Exception as template_error:
                self.logger.warning(f"Email template not found, using fallback: {template_error}")

            success = self.send_email(email, subject, html_content)
            if success:
                self.logger.info(f"OTP email sent successfully to {email}")
            else:
                self.logger.error(f"Failed to send OTP email to {email}")

            return success

        except Exception as e:
            self.logger.error(f"Error in send_otp_email: {str(e)}")
            return False

    def _create_otp_email_fallback(self, otp_code):
        """Create fallback OTP email content"""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px 20px; text-align: center; }}
                .content {{ padding: 30px; }}
                .otp-code {{ font-size: 42px; font-weight: bold; color: #667eea; text-align: center; margin: 30px 0; padding: 20px; background: #f8f9fa; border-radius: 8px; letter-spacing: 8px; }}
                .info-box {{ background: #e7f3ff; border-left: 4px solid #667eea; padding: 15px; margin: 20px 0; border-radius: 4px; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
                .button {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 12px 30px;
                    text-decoration: none;
                    border-radius: 5px;
                    display: inline-block;
                    margin: 10px 0;
                    font-weight: bold;
                    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔐 FiscalFlow</h1>
                    <p>Financial Management Simplified</p>
                </div>
                <div class="content">
                    <h2>Email Verification Required</h2>
                    <p>Hello,</p>
                    <p>Thank you for choosing FiscalFlow! To complete your registration and secure your account, please use the following verification code:</p>

                    <div class="otp-code">{otp_code}</div>

                    <div class="info-box">
                        <strong>Important:</strong>
                        <ul>
                            <li>This code will expire in 10 minutes</li>
                            <li>Do not share this code with anyone</li>
                            <li>If you didn't request this verification, please ignore this email</li>
                        </ul>
                    </div>

                    <p>Enter this code in the verification page to activate your account and start managing your finances with FiscalFlow.</p>

                    <p>Need help? Contact our support team at <a href="mailto:support@fiscalflow.com">support@fiscalflow.com</a></p>
                </div>
                <div class="footer">
                    <p>&copy; 2024 FiscalFlow. All rights reserved.</p>
                    <p>This is an automated message, please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>
        """

    def send_password_reset(self, email, token):
        """Send password reset email with proper button styling - SUPPORTS LOCAL AND PRODUCTION"""
        try:
            # Determine base URL based on environment
            if self.app.config.get('DEBUG') or self.app.config.get('ENV') == 'development':
                # Local development URL
                base_url = "http://localhost:5000"
            else:
                # Production URL
                base_url = "https://fiscal01.pythonanywhere.com"

            reset_url = f"{base_url}/reset-password/{token}"

            self.logger.info(f"Generating password reset for: {email}")
            self.logger.info(f"Reset URL: {reset_url}")
            self.logger.info(f"Token: {token}")

            subject = "Reset Your FiscalFlow Password"

            # Try template first
            try:
                html_content = render_template('email/reset_password.html',
                                             reset_url=reset_url,
                                             base_url=base_url)
                self.logger.info("Reset password template rendered successfully")

            except Exception as template_error:
                self.logger.warning(f"Reset password template not found, using fallback: {template_error}")
                html_content = self._create_password_reset_fallback(reset_url, base_url)

            # Log the HTML content for debugging
            self.logger.info(f"Email HTML content length: {len(html_content)}")

            success = self.send_email(email, subject, html_content)
            if success:
                self.logger.info(f"Password reset email sent successfully to {email}")
            else:
                self.logger.error(f"Failed to send password reset email to {email}")
            return success

        except Exception as e:
            self.logger.error(f"Error in send_password_reset: {str(e)}")
            self.logger.error(traceback.format_exc())
            return False

    def _create_password_reset_fallback(self, reset_url, base_url=None):
        """Create fallback password reset email content with proper button styling"""
        # Ensure we have the full URL
        if not reset_url.startswith(('http://', 'https://')):
            if not base_url:
                base_url = self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000')
            reset_url = f"{base_url}{reset_url}"

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px 20px; text-align: center; }}
                .content {{ padding: 30px; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
                .url-backup {{
                    word-break: break-all;
                    background: #f8f9fa;
                    padding: 12px;
                    border-radius: 5px;
                    margin: 10px 0;
                    font-family: 'Courier New', monospace;
                    font-size: 14px;
                    border: 1px solid #e9ecef;
                }}
                .reset-button {{
                    background-color: #667eea;
                    border: 1px solid #667eea;
                    border-radius: 8px;
                    color: #ffffff;
                    display: inline-block;
                    font-family: Arial, sans-serif;
                    font-size: 16px;
                    font-weight: bold;
                    line-height: 1;
                    padding: 15px 35px;
                    text-decoration: none;
                    text-align: center;
                    margin: 20px 0;
                }}
                .reset-button:hover {{
                    background-color: #5a6fd8;
                    border-color: #5a6fd8;
                }}
                .info-box {{
                    background: #e7f3ff;
                    border-left: 4px solid #667eea;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 4px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔒 FiscalFlow</h1>
                    <p>Password Reset Request</p>
                </div>
                <div class="content">
                    <h2>Reset Your Password</h2>
                    <p>We received a request to reset your FiscalFlow account password.</p>

                    <!-- Main Reset Button -->
                    <div style="text-align: center;">
                        <a href="{reset_url}" class="reset-button" target="_blank">
                            Reset Your Password
                        </a>
                    </div>

                    <p>If the button doesn't work, copy and paste this link into your browser:</p>
                    <div class="url-backup">{reset_url}</div>

                    <div class="info-box">
                        <strong>Security Notice:</strong>
                        <ul>
                            <li>This link will expire in 1 hour</li>
                            <li>If you didn't request this reset, please ignore this email</li>
                            <li>Your account security is important to us</li>
                        </ul>
                    </div>

                    <p>For security reasons, this link can only be used once and will expire automatically.</p>

                    <p>Need help? Contact our support team at <a href="mailto:support@fiscalflow.com">support@fiscalflow.com</a></p>
                </div>
                <div class="footer">
                    <p>&copy; 2024 FiscalFlow. All rights reserved.</p>
                    <p>This is an automated security message.</p>
                </div>
            </div>
        </body>
        </html>
        """

    def send_transaction_update(self, email, customer_name, transaction_type, amount, balance):
        """Send transaction update notification to customer"""
        try:
            subject = f"Transaction Update - {customer_name}"

            # Fallback HTML content
            html_content = self._create_transaction_update_fallback(customer_name, transaction_type, amount, balance)

            # Try template
            try:
                html_content = render_template('email/transaction_update.html',
                                            customer_name=customer_name,
                                            transaction_type=transaction_type,
                                            amount=amount,
                                            balance=balance,
                                            app_url=self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000'))
            except Exception as template_error:
                self.logger.warning(f"Transaction update template not found: {template_error}")

            success = self.send_email(email, subject, html_content)
            if success:
                self.logger.info(f"Transaction update sent to {email}")
            return success

        except Exception as e:
            self.logger.error(f"Error in send_transaction_update: {str(e)}")
            return False

    def _create_transaction_update_fallback(self, customer_name, transaction_type, amount, balance):
        """Create fallback transaction update email content"""
        transaction_type_display = transaction_type.replace('_', ' ').title()
        amount_formatted = f"${amount:,.2f}"
        balance_formatted = f"${balance:,.2f}"

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #20bf6b 0%, #01baef 100%); color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 25px; }}
                .transaction-details {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0; }}
                .amount {{ font-size: 24px; font-weight: bold; color: #20bf6b; text-align: center; margin: 15px 0; }}
                .balance {{ font-size: 20px; color: #2d3436; text-align: center; margin: 15px 0; }}
                .footer {{ background: #f8f9fa; padding: 15px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
                .view-details-button {{
                    background: linear-gradient(135deg, #20bf6b 0%, #01baef 100%);
                    color: white;
                    padding: 12px 25px;
                    text-decoration: none;
                    border-radius: 6px;
                    display: inline-block;
                    margin: 15px 0;
                    font-weight: bold;
                    box-shadow: 0 4px 15px rgba(32, 191, 107, 0.3);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>💳 FiscalFlow</h1>
                    <p>Transaction Notification</p>
                </div>
                <div class="content">
                    <h2>Hello {customer_name},</h2>
                    <p>A new transaction has been recorded in your account:</p>

                    <div class="transaction-details">
                        <div class="amount">{transaction_type_display}: {amount_formatted}</div>
                        <div class="balance">Current Balance: {balance_formatted}</div>
                    </div>

                    <div style="text-align: center;">
                        <a href="{self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000')}/dashboard" class="view-details-button">View Account Details</a>
                    </div>

                    <p>This is an automated notification from FiscalFlow.</p>
                    <p>If you have any questions about this transaction, please contact your account manager.</p>
                </div>
                <div class="footer">
                    <p>&copy; 2024 FiscalFlow. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """

    def send_balance_update(self, email, customer_name, balance):
        """Send balance update notification to customer"""
        try:
            subject = f"Account Balance Update - {customer_name}"

            # Fallback HTML content
            html_content = self._create_balance_update_fallback(customer_name, balance)

            # Try template
            try:
                html_content = render_template('email/balance_update.html',
                                            customer_name=customer_name,
                                            balance=balance,
                                            app_url=self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000'))
            except Exception as template_error:
                self.logger.warning(f"Balance update template not found: {template_error}")

            success = self.send_email(email, subject, html_content)
            if success:
                self.logger.info(f"Balance update sent to {email}")
            return success

        except Exception as e:
            self.logger.error(f"Error in send_balance_update: {str(e)}")
            return False

    def _create_balance_update_fallback(self, customer_name, balance):
        """Create fallback balance update email content"""
        balance_formatted = f"${balance:,.2f}"

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #a55eea 0%, #8854d0 100%); color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 25px; }}
                .balance-display {{ background: #f8f9fa; padding: 25px; border-radius: 8px; margin: 20px 0; text-align: center; }}
                .balance-amount {{ font-size: 32px; font-weight: bold; color: #a55eea; margin: 10px 0; }}
                .footer {{ background: #f8f9fa; padding: 15px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
                .view-account-button {{
                    background: linear-gradient(135deg, #a55eea 0%, #8854d0 100%);
                    color: white;
                    padding: 12px 25px;
                    text-decoration: none;
                    border-radius: 6px;
                    display: inline-block;
                    margin: 15px 0;
                    font-weight: bold;
                    box-shadow: 0 4px 15px rgba(165, 94, 234, 0.3);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>💰 FiscalFlow</h1>
                    <p>Balance Update</p>
                </div>
                <div class="content">
                    <h2>Hello {customer_name},</h2>
                    <p>Your account balance has been updated:</p>

                    <div class="balance-display">
                        <div style="font-size: 16px; color: #666;">Current Balance</div>
                        <div class="balance-amount">{balance_formatted}</div>
                    </div>

                    <div style="text-align: center;">
                        <a href="{self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000')}/dashboard" class="view-account-button">View Account</a>
                    </div>

                    <p>This is an automated notification from FiscalFlow.</p>
                    <p>For detailed transaction history, please log in to your account or contact your account manager.</p>
                </div>
                <div class="footer">
                    <p>&copy; 2024 FiscalFlow. All rights reserved.</p>
                </div>
            </div>
        </body>
        </html>
        """

    def send_welcome_email(self, email, customer_name):
        """Send welcome email to new customers"""
        try:
            subject = "Welcome to FiscalFlow!"

            html_content = self._create_welcome_email_fallback(customer_name)

            # Try template
            try:
                html_content = render_template('email/welcome.html',
                                            customer_name=customer_name,
                                            app_url=self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000'))
            except Exception as template_error:
                self.logger.warning(f"Welcome email template not found: {template_error}")

            success = self.send_email(email, subject, html_content)
            if success:
                self.logger.info(f"Welcome email sent to {email}")
            return success

        except Exception as e:
            self.logger.error(f"Error in send_welcome_email: {str(e)}")
            return False

    def _create_welcome_email_fallback(self, customer_name):
        """Create fallback welcome email content"""
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; margin: 0; padding: 20px; }}
                .container {{ max-width: 600px; margin: 0 auto; background: #ffffff; border: 1px solid #ddd; border-radius: 8px; overflow: hidden; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px 20px; text-align: center; }}
                .content {{ padding: 30px; }}
                .feature {{ display: flex; align-items: center; margin: 15px 0; }}
                .feature-icon {{ font-size: 24px; margin-right: 15px; }}
                .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 12px; border-top: 1px solid #ddd; }}
                .get-started-button {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    color: white;
                    padding: 15px 35px;
                    text-decoration: none;
                    border-radius: 8px;
                    display: inline-block;
                    margin: 20px 0;
                    font-size: 16px;
                    font-weight: bold;
                    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🎉 Welcome to FiscalFlow!</h1>
                    <p>Financial Management Simplified</p>
                </div>
                <div class="content">
                    <h2>Hello {customer_name},</h2>
                    <p>Welcome to FiscalFlow! We're excited to help you manage your finances effectively.</p>

                    <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <h3 style="margin-top: 0;">What you can do with FiscalFlow:</h3>
                        <div class="feature">
                            <div class="feature-icon">💳</div>
                            <div>Track income and expenses</div>
                        </div>
                        <div class="feature">
                            <div class="feature-icon">📊</div>
                            <div>View detailed financial reports</div>
                        </div>
                        <div class="feature">
                            <div class="feature-icon">🔔</div>
                            <div>Receive transaction notifications</div>
                        </div>
                        <div class="feature">
                            <div class="feature-icon">📱</div>
                            <div>Access your account anywhere</div>
                        </div>
                    </div>

                    <div style="text-align: center;">
                        <a href="{self.app.config.get('FISCALFLOW_URL', 'http://localhost:5000')}/dashboard" class="get-started-button">Get Started</a>
                    </div>

                    <p>Need help getting started? Our support team is here to assist you.</p>
                    <p>Best regards,<br>The FiscalFlow Team</p>
                </div>
                <div class="footer">
                    <p>&copy; 2024 FiscalFlow. All rights reserved.</p>
                    <p>This is an automated welcome message.</p>
                </div>
            </div>
        </body>
        </html>
        """

    def send_email(self, to, subject, html_body):
        """Send email with comprehensive error handling and connection testing"""
        try:
            # Validate email configuration
            required_config = ['MAIL_SERVER', 'MAIL_PORT', 'MAIL_USERNAME', 'MAIL_PASSWORD']
            for config_key in required_config:
                if not self.app.config.get(config_key):
                    self.logger.error(f"Missing email configuration: {config_key}")
                    return False

            # Test connection before sending
            if not self._test_email_connection():
                self.logger.error("Email connection test failed")
                return False

            # Create message
            msg = Message(
                subject=subject,
                recipients=[to],
                html=html_body,
                sender=self.app.config['MAIL_DEFAULT_SENDER'],
                reply_to=self.app.config.get('MAIL_DEFAULT_SENDER')
            )

            # Send email
            self.mail.send(msg)
            self.logger.info(f"Email sent successfully to {to}")
            return True

        except smtplib.SMTPAuthenticationError as e:
            self.logger.error(f"SMTP Authentication failed: {e}")
            self.logger.error("Please check:")
            self.logger.error("1. Email username and password are correct")
            self.logger.error("2. For Gmail: Enable 2FA and use App Password")
            self.logger.error("3. For other providers: Check SMTP settings")
            return False

        except smtplib.SMTPConnectError as e:
            self.logger.error(f"SMTP Connection failed: {e}")
            self.logger.error("Please check:")
            self.logger.error("1. MAIL_SERVER and MAIL_PORT are correct")
            self.logger.error("2. Internet connection is available")
            self.logger.error("3. Firewall allows SMTP connections")
            return False

        except smtplib.SMTPSenderRefused as e:
            self.logger.error(f"SMTP Sender refused: {e}")
            return False

        except smtplib.SMTPRecipientsRefused as e:
            self.logger.error(f"SMTP Recipient refused: {e}")
            return False

        except smtplib.SMTPException as e:
            self.logger.error(f"SMTP error occurred: {e}")
            return False

        except Exception as e:
            self.logger.error(f"Unexpected error sending email: {e}")
            return False

    def _test_email_connection(self):
        """Test SMTP connection without sending email"""
        try:
            server = smtplib.SMTP(self.app.config['MAIL_SERVER'], self.app.config['MAIL_PORT'])
            server.ehlo()

            if self.app.config.get('MAIL_USE_TLS'):
                server.starttls()
                server.ehlo()

            server.login(self.app.config['MAIL_USERNAME'], self.app.config['MAIL_PASSWORD'])
            server.quit()

            self.logger.info("Email connection test successful")
            return True

        except Exception as e:
            self.logger.error(f"Email connection test failed: {e}")
            return False

class FileUtils:
    @staticmethod
    def allowed_file(filename, allowed_extensions=None):
        """Check if file extension is allowed"""
        if allowed_extensions is None:
            allowed_extensions = {'png', 'jpg', 'jpeg', 'pdf', 'doc', 'docx'}

        if '.' not in filename:
            return False

        extension = filename.rsplit('.', 1)[1].lower()
        return extension in allowed_extensions

    @staticmethod
    def save_uploaded_file(file, upload_folder, max_size=(800, 600)):
        """Save uploaded file with validation and processing"""
        try:
            if not file or not FileUtils.allowed_file(file.filename):
                return None

            # Generate secure filename
            original_name = secure_filename(file.filename)
            file_extension = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
            random_prefix = secrets.token_hex(8)
            filename = f"{random_prefix}_{original_name}"
            filepath = os.path.join(upload_folder, filename)

            # Create upload folder if it doesn't exist
            os.makedirs(upload_folder, exist_ok=True)

            # Save file
            file.save(filepath)

            # Process image files (resize)
            if file_extension in ['png', 'jpg', 'jpeg']:
                try:
                    with Image.open(filepath) as img:
                        # Convert to RGB if necessary
                        if img.mode in ('RGBA', 'P'):
                            img = img.convert('RGB')

                        # Resize if larger than max_size
                        if img.size[0] > max_size[0] or img.size[1] > max_size[1]:
                            img.thumbnail(max_size, Image.Resampling.LANCZOS)
                            img.save(filepath, optimize=True, quality=85)
                except Exception as img_error:
                    logger.warning(f"Image processing failed, keeping original: {img_error}")

            logger.info(f"File saved successfully: {filename}")
            return filename

        except Exception as e:
            logger.error(f"Error saving file: {e}")
            return None

    @staticmethod
    def delete_file(filename, upload_folder):
        """Delete uploaded file"""
        try:
            if filename:
                filepath = os.path.join(upload_folder, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    logger.info(f"File deleted: {filename}")
                    return True
            return False
        except Exception as e:
            logger.error(f"Error deleting file: {e}")
            return False

class OTPUtils:
    @staticmethod
    def generate_otp(length=6):
        """Generate a secure OTP code"""
        return ''.join(secrets.choice('0123456789') for _ in range(length))

    @staticmethod
    def is_otp_valid(otp_record):
        """Check if OTP is valid and not expired"""
        if not otp_record:
            return False
        if otp_record.is_used:
            return False
        if otp_record.expires_at <= datetime.utcnow():
            return False
        return True

    @staticmethod
    def cleanup_expired_otps(db_session, OTPModel):
        """Clean up expired OTP records"""
        try:
            expired_count = OTPModel.query.filter(
                OTPModel.expires_at <= datetime.utcnow()
            ).delete()
            db_session.commit()
            if expired_count > 0:
                logger.info(f"Cleaned up {expired_count} expired OTP records")
            return expired_count
        except Exception as e:
            logger.error(f"Error cleaning up expired OTPs: {e}")
            db_session.rollback()
            return 0

class ReportUtils:
    """
    Utility class for generating PDF and Excel reports.
    Works on PythonAnywhere and all modern WeasyPrint versions.
    """

    @staticmethod
    def generate_pdf_report(customer, transactions):
        """
        Generate a PDF report for customer transactions using WeasyPrint.
        Returns PDF bytes.
        """
        try:
            logger.info(f"[PDF] Generating PDF for: {customer.name}")

            # Import WeasyPrint
            from weasyprint import HTML
            
            # Get HTML content
            html_content = ReportUtils._get_pdf_html_content(customer, transactions)

            if not html_content or not html_content.strip():
                logger.error("[PDF] HTML content is empty")
                raise Exception("Failed to generate HTML content for PDF")

            logger.info(f"[PDF] HTML content ready ({len(html_content)} characters)")

            # Create HTML object and generate PDF
            html_obj = HTML(string=html_content)
            
            logger.info("[PDF] Converting HTML → PDF…")
            pdf_bytes = html_obj.write_pdf()

            if not pdf_bytes:
                raise Exception("WeasyPrint returned empty PDF")

            logger.info(f"[PDF] PDF generated successfully ({len(pdf_bytes)} bytes)")
            return pdf_bytes

        except ImportError as e:
            logger.error(f"[PDF] WeasyPrint not installed: {e}")
            raise Exception("PDF generation requires WeasyPrint. Please install it with: pip install weasyprint")
        
        except Exception as e:
            logger.error(f"[PDF] PDF generation failed: {e}")
            logger.error(traceback.format_exc())
            raise Exception(f"PDF generation failed: {e}")

    @staticmethod
    def _get_pdf_html_content(customer, transactions):
        """Get HTML content for PDF generation with fallbacks"""
        try:
            # Try template first
            html_content = render_template(
                "pdf_template.html",
                customer=customer,
                transactions=transactions,
                generated_at=datetime.utcnow()
            )
            logger.info("[PDF] Successfully rendered template")
            return html_content
            
        except Exception as template_error:
            logger.warning(f"[PDF] Template failed, using fallback: {template_error}")
            return ReportUtils._create_pdf_fallback_template(customer, transactions)

    @staticmethod
    def _create_pdf_fallback_template(customer, transactions):
        """
        Create fallback PDF template when the template file is not found.
        """
        logger.info("[PDF] Using fallback PDF template")

        # Calculate totals
        cash_in_total = sum(t.amount for t in transactions if t.type == 'cash_in')
        cash_out_total = sum(t.amount for t in transactions if t.type == 'cash_out')

        # Format balance display
        balance_class = "positive-balance" if customer.current_balance >= 0 else "negative-balance"

        # Generate transactions HTML
        transactions_html = ""
        if transactions:
            for transaction in transactions:
                transaction_type_class = "cash-in" if transaction.type == 'cash_in' else "cash-out"
                transaction_type_display = transaction.type.replace('_', ' ').title()

                transactions_html += f"""
                <tr>
                    <td>{transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A'}</td>
                    <td class="transaction-type">{transaction_type_display}</td>
                    <td class="amount-cell {transaction_type_class}">₹{transaction.amount:.2f}</td>
                    <td>{transaction.category or 'General'}</td>
                    <td>{transaction.notes or 'No description'}</td>
                </tr>
                """
        else:
            transactions_html = """
            <tr>
                <td colspan="5" style="text-align: center; padding: 40px; color: #1565c0; font-style: italic;">
                    <h3>No Transaction Records</h3>
                    <p>No transaction history available for this customer.</p>
                </td>
            </tr>
            """

        current_time = datetime.utcnow()

        # Create transactions table
        if transactions:
            transactions_table = f"""
            <table class='transaction-table'>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Transaction Type</th>
                        <th>Amount (₹)</th>
                        <th>Category</th>
                        <th>Description</th>
                    </tr>
                </thead>
                <tbody>{transactions_html}</tbody>
            </table>
            """
        else:
            transactions_table = """
            <div class="no-data">
                <h3>No Transaction Records</h3>
                <p>No transaction history available for this customer.</p>
            </div>
            """

        credit_limit_html = f"<p><strong>Credit Limit:</strong> ₹{customer.credit_limit:.2f}</p>" if customer.credit_limit and customer.credit_limit > 0 else ""

        return f"""
        <!DOCTYPE html>
        <html>
          <head>
            <meta charset="UTF-8" />
            <title>Transaction Report - {customer.name}</title>
            <style>
              @page {{
                size: A4;
                margin: 2cm;
                @top-center {{
                  content: "FiscalFlow - Transaction Report";
                  font-size: 12px;
                  font-family: Arial, sans-serif;
                  color: #1565c0;
                }}
                @bottom-center {{
                  content: "Page " counter(page) " of " counter(pages);
                  font-size: 10px;
                  font-family: Arial, sans-serif;
                  color: #1565c0;
                }}
              }}

              body {{
                font-family: "Arial", sans-serif;
                margin: 0;
                padding: 0;
                font-size: 12px;
                line-height: 1.4;
                color: #1565c0;
                background-color: #ffffff;
              }}

              .header {{
                text-align: center;
                margin-bottom: 30px;
                border-bottom: 3px solid #1565c0;
                padding-bottom: 20px;
              }}

              .header h1 {{
                font-size: 24px;
                margin: 0 0 10px 0;
                color: #1565c0;
                font-weight: bold;
              }}

              .header h2 {{
                font-size: 18px;
                margin: 0;
                color: #1976d2;
                font-weight: normal;
              }}

              .customer-info {{
                margin-bottom: 25px;
                background: #f8fbff;
                padding: 20px;
                border-radius: 8px;
                border: 2px solid #e3f2fd;
              }}

              .customer-info p {{
                margin: 8px 0;
                font-size: 13px;
                color: #1565c0;
              }}

              .customer-info strong {{
                color: #0d47a1;
                font-weight: bold;
              }}

              .balance-highlight {{
                font-size: 14px;
                font-weight: bold;
                color: #1565c0;
                background-color: #e3f2fd;
                padding: 8px 12px;
                border-radius: 6px;
                display: inline-block;
                margin-top: 5px;
                border: 1px solid #bbdefb;
              }}

              .transaction-table {{
                width: 100%;
                border-collapse: collapse;
                margin-top: 20px;
                font-size: 11px;
                border: 2px solid #1565c0;
                border-radius: 8px;
                overflow: hidden;
              }}

              .transaction-table th {{
                background: #1565c0;
                color: white;
                font-weight: bold;
                text-transform: uppercase;
                font-size: 10px;
                letter-spacing: 0.5px;
                padding: 12px 8px;
                text-align: left;
                border: 1px solid #0d47a1;
              }}

              .transaction-table td {{
                padding: 10px 8px;
                border: 1px solid #bbdefb;
                color: #1565c0;
              }}

              .transaction-table tr:nth-child(even) {{
                background-color: #f8fbff;
              }}

              .transaction-table tr:nth-child(odd) {{
                background-color: #ffffff;
              }}

              .summary {{
                margin-top: 30px;
                padding: 20px;
                background: #f8fbff;
                border: 2px solid #e3f2fd;
                border-radius: 8px;
              }}

              .summary p {{
                margin: 8px 0;
                font-size: 13px;
                color: #1565c0;
              }}

              .summary strong {{
                color: #0d47a1;
              }}

              .cash-in {{
                color: #1565c0;
                font-weight: bold;
                background-color: #e8f5e8;
                padding: 4px 8px;
                border-radius: 4px;
                border: 1px solid #c8e6c9;
              }}

              .cash-out {{
                color: #1565c0;
                font-weight: bold;
                background-color: #ffebee;
                padding: 4px 8px;
                border-radius: 4px;
                border: 1px solid #ffcdd2;
              }}

              .no-data {{
                text-align: center;
                padding: 40px;
                color: #1565c0;
                font-style: italic;
                background-color: #f8fbff;
                border: 2px dashed #bbdefb;
                border-radius: 8px;
                margin: 20px 0;
              }}

              .no-data h3 {{
                color: #1565c0;
                margin-bottom: 10px;
              }}

              .report-meta {{
                font-size: 10px;
                color: #1565c0;
                text-align: center;
                margin-top: 30px;
                padding: 15px;
                background-color: #f8fbff;
                border-radius: 8px;
                border-top: 2px solid #bbdefb;
              }}

              .transaction-type {{
                text-transform: capitalize;
                font-weight: bold;
              }}

              .section-title {{
                font-size: 16px;
                font-weight: bold;
                color: #1565c0;
                margin: 25px 0 15px 0;
                padding-bottom: 8px;
                border-bottom: 2px solid #e3f2fd;
              }}

              .positive-balance {{
                color: #1565c0;
                background-color: #e8f5e8;
              }}

              .negative-balance {{
                color: #1565c0;
                background-color: #ffebee;
              }}

              .amount-cell {{
                font-weight: bold;
                text-align: right;
              }}

              .footer-note {{
                text-align: center;
                font-size: 10px;
                color: #1976d2;
                margin-top: 20px;
                font-style: italic;
              }}
            </style>
          </head>
          <body>
            <div class="header">
              <h1>FISCALFLOW TRANSACTION REPORT</h1>
              <h2>{customer.name}</h2>
            </div>

            <div class="customer-info">
              <p><strong>Customer Name:</strong> {customer.name}</p>
              <p><strong>Email Address:</strong> {customer.email or 'Not Provided'}</p>
              <p><strong>Phone Number:</strong> {customer.phone or 'Not Provided'}</p>
              <p>
                <strong>Current Balance:</strong>
                <span class="balance-highlight {balance_class}">
                  ₹{customer.current_balance:.2f}
                </span>
              </p>
              {credit_limit_html}
            </div>

            <div class="section-title">TRANSACTION HISTORY</div>

            {transactions_table}

            <div class="section-title">REPORT SUMMARY</div>

            <div class="summary">
              <p><strong>Total Number of Transactions:</strong> {len(transactions)}</p>
              <p><strong>Total Cash In:</strong> <span class="cash-in">₹{cash_in_total:.2f}</span></p>
              <p><strong>Total Cash Out:</strong> <span class="cash-out">₹{cash_out_total:.2f}</span></p>
              <p><strong>Reporting Period:</strong> Complete Transaction History</p>
              <p><strong>Account Status:</strong> {'Active' if transactions else 'No Activity'}</p>
            </div>

            <div class="footer-note">
              This report contains confidential financial information. Please handle
              with appropriate security measures.
            </div>

            <div class="report-meta">
              <p><strong>Report Generated:</strong> {current_time.strftime('%Y-%m-%d at %H:%M UTC')}</p>
              <p><strong>System:</strong> FiscalFlow Financial Management</p>
              <p><strong>Currency:</strong> Indian Rupee (₹)</p>
              <p><strong>Report ID:</strong> {customer.id}_{current_time.strftime('%Y%m%d%H%M')}</p>
            </div>
          </body>
        </html>
        """

    @staticmethod
    def generate_excel_report(customer, transactions):
        """
        Generate Excel (.xlsx) report with the following format:
        - Customer name and balance at top
        - Table columns: Date, Time, Note, Cash In, Cash Out, Balance
        - Separate columns for cash in and cash out amounts
        - Running balance calculation
        Returns Excel bytes.
        """
        try:
            logger.info(f"[Excel] Generating Excel for: {customer.name}")

            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO

            wb = Workbook()
            ws = wb.active
            ws.title = f"{customer.name[:25]} Transactions"  # Limit sheet name length

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            title_font = Font(bold=True, size=16, color="1565C0")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))

            # Title Section
            ws.merge_cells('A1:F1')
            ws['A1'] = f"TRANSACTION REPORT - {customer.name.upper()}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # Customer Information Section
            ws.merge_cells('A2:F2')
            ws['A2'] = "Customer Information"
            ws['A2'].font = Font(bold=True, size=12, color="1565C0")
            ws['A2'].alignment = center_align

            ws['A3'] = "Customer Name:"
            ws['A3'].font = bold_font
            ws['B3'] = customer.name

            # Add contact information if available
            if customer.email or customer.phone:
                row = 5
                if customer.email:
                    ws[f'A{row}'] = "Email:"
                    ws[f'A{row}'].font = bold_font
                    ws[f'B{row}'] = customer.email or "Not Provided"
                    row += 1
                if customer.phone:
                    ws[f'A{row}'] = "Phone:"
                    ws[f'A{row}'].font = bold_font
                    ws[f'B{row}'] = customer.phone or "Not Provided"
                    row += 1

            # Add spacing
            empty_row = 7
            ws[f'A{empty_row}'] = ""

            # Table Header - Start from row 8
            headers = ["Date", "Time", "Note", "Cash In (₹)", "Cash Out (₹)", "Balance (₹)"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=8, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # Transaction Rows - Start from row 9
            row_num = 9
            running_balance = 0
            
            # Sort transactions by date for proper balance calculation
            sorted_transactions = sorted(transactions, key=lambda x: x.date if x.date else datetime.min)
            
            for transaction in sorted_transactions:
                # Date
                date_cell = ws.cell(row=row_num, column=1, 
                                  value=transaction.date.strftime('%Y-%m-%d') if transaction.date else 'N/A')
                date_cell.border = thin_border
                date_cell.alignment = left_align

                # Time
                time_cell = ws.cell(row=row_num, column=2,
                                  value=transaction.date.strftime('%H:%M:%S') if transaction.date else 'N/A')
                time_cell.border = thin_border
                time_cell.alignment = left_align

                # Note/Description
                note_cell = ws.cell(row=row_num, column=3, 
                                  value=transaction.notes or transaction.category or "Transaction")
                note_cell.border = thin_border
                note_cell.alignment = left_align

                # Cash In amount
                cash_in_cell = ws.cell(row=row_num, column=4)
                cash_in_cell.border = thin_border
                cash_in_cell.alignment = right_align
                cash_in_cell.number_format = '#,##0.00'

                # Cash Out amount
                cash_out_cell = ws.cell(row=row_num, column=5)
                cash_out_cell.border = thin_border
                cash_out_cell.alignment = right_align
                cash_out_cell.number_format = '#,##0.00'

                # Update running balance and set cash in/out values
                if transaction.type == 'cash_in':
                    cash_in_cell.value = transaction.amount
                    running_balance += transaction.amount
                else:  # cash_out
                    cash_out_cell.value = transaction.amount
                    running_balance -= transaction.amount

                # Balance
                balance_cell = ws.cell(row=row_num, column=6, value=running_balance)
                balance_cell.border = thin_border
                balance_cell.alignment = right_align
                balance_cell.number_format = '#,##0.00'
                balance_cell.font = bold_font

                row_num += 1

            # If no transactions, show message
            if not transactions:
                ws.merge_cells(f'A{row_num}:F{row_num}')
                no_data_cell = ws.cell(row=row_num, column=1, value="No transaction records available")
                no_data_cell.font = Font(italic=True, color="1565C0", size=12)
                no_data_cell.alignment = center_align
                no_data_cell.border = thin_border

            # Auto-fit columns with some padding
            column_widths = {
                'A': 12,  # Date
                'B': 10,  # Time
                'C': 30,  # Note
                'D': 12,  # Cash In
                'E': 12,  # Cash Out
                'F': 15   # Balance
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()

            logger.info(f"[Excel] Excel generated successfully ({len(excel_bytes)} bytes)")
            return excel_bytes

        except ImportError as e:
            logger.error(f"[Excel] OpenPyXL not installed: {e}")
            raise Exception("Excel generation requires OpenPyXL. Please install it with: pip install openpyxl")
        
        except Exception as e:
            logger.error(f"[Excel] Excel report generation failed: {e}")
            logger.error(traceback.format_exc())
            raise Exception(f"Excel generation error: {e}")